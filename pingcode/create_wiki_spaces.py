"""
PingCode 知识管理空间批量创建脚本

功能:
  1. 读取 pingcode/导入表.xlsx
  2. 通过企业令牌认证
  3. 批量创建知识管理空间 (scope_type=organization)
  4. 为每个空间设置管理员

Excel 格式:
  第1列: 空间名称
  第2列: 空间标识
  第3列: 管理员姓名 (多个以逗号分隔)

使用前修改下方 CONFIG 区域的配置项。

依赖: pip install requests openpyxl
"""

import os
import re
import sys
import requests
from openpyxl import load_workbook

# ============================================================
# CONFIG - 使用前请修改以下配置
# ============================================================

# 企业 Client ID
CLIENT_ID = "ePrgDnMsGmZT"

# 企业 Client Secret
CLIENT_SECRET = "SWrynTvzilneobmtruhuCHRI"

# PingCode API 地址
# SaaS 部署: https://open.pingcode.com
# 私有部署: https://pingcode.yourcompany.com
BASE_URL = "https://open.pingcode.com"

# 私有部署设置为 True，SaaS 部署设置为 False
IS_PRIVATE = False

# 认证接口路径
AUTH_TOKEN_PATH = "/v1/auth/token"

# 用户查询接口路径
USERS_PATH = "/v1/directory/users"

# 知识空间接口路径
WIKI_SPACES_PATH = "/v1/wiki/spaces"

# 知识空间成员接口路径模板 (使用 {space_id} 占位)
WIKI_MEMBERS_PATH = "/v1/wiki/spaces/{space_id}/members"

# 每个 API 请求的超时时间(秒)
REQUEST_TIMEOUT = 30

# admin 角色 ID（PingCode 内置管理员角色，可通过 GET /v1/directory/roles 查询）
ADMIN_ROLE_ID = "100000000000000000000001"

# 空间 scope_type，organization 表示组织级别可见
SCOPE_TYPE = "organization"

# ============================================================
# END CONFIG
# ============================================================


def api_path(path: str) -> str:
    """根据是否为私有部署，返回带正确前缀的完整 URL"""
    prefix = "/open" if IS_PRIVATE else ""
    return f"{BASE_URL}{prefix}{path}"


def get_enterprise_token() -> str:
    """通过 client_credentials 获取企业 access_token"""
    url = api_path(AUTH_TOKEN_PATH)
    params = {
        "grant_type": "client_credentials",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
    }
    resp = requests.get(url, params=params, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    data = resp.json()
    access_token = data.get("access_token")
    if not access_token:
        raise RuntimeError(f"获取企业令牌失败，响应: {data}")
    return access_token


def find_user_by_name(token: str, name: str) -> dict | None:
    """按姓名查找用户，返回 {'id': ..., 'display_name': ..., 'email': ...}"""
    url = api_path(USERS_PATH)
    resp = requests.get(
        url,
        params={"keywords": name, "page_index": 0, "page_size": 10},
        headers={"Authorization": f"Bearer {token}"},
        timeout=REQUEST_TIMEOUT,
    )
    resp.raise_for_status()
    users = resp.json().get("values", [])
    if not users:
        return None
    for user in users:
        if user.get("display_name") == name.strip():
            return {"id": user["id"], "display_name": user["display_name"], "email": user.get("email", "")}
    user = users[0]
    return {"id": user["id"], "display_name": user["display_name"], "email": user.get("email", "")}


def create_wiki_space(token: str, name: str, identifier: str) -> dict:
    """创建知识管理空间，返回响应数据"""
    url = api_path(WIKI_SPACES_PATH)
    body = {
        "name": name,
        "identifier": identifier,
        "scope_type": SCOPE_TYPE,
    }
    resp = requests.post(url, json=body, headers={"Authorization": f"Bearer {token}"}, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    return resp.json()


def add_member(token: str, space_id: str, user_id: str, role_id: str = ADMIN_ROLE_ID) -> dict:
    """将用户添加为空间成员"""
    url = api_path(WIKI_MEMBERS_PATH.format(space_id=space_id))
    body = {"member": {"id": user_id, "type": "user"}, "role_id": role_id}
    resp = requests.post(url, json=body, headers={"Authorization": f"Bearer {token}"}, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    return resp.json()


def read_excel(filepath: str) -> list[tuple[str, str, list[str]]]:
    """
    读取 Excel 文件，自动跳过表头行
    返回 [(空间名称, 空间标识, [管理员名1, 管理员名2, ...]), ...]
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"文件不存在: {filepath}")

    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        name = str(row[0]).strip() if row[0] is not None else ""
        identifier = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        admins_str = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""

        if not name or not identifier:
            print(f"  ⚠ 第{row_idx}行数据不完整（名称={name!r}, 标识={identifier!r}），跳过")
            continue

        if name in ("空间名称", "Name") and identifier in ("空间标识", "Identifier"):
            print(f"  ℹ 第{row_idx}行为表头，自动跳过")
            continue

        admin_names = [a.strip() for a in re.split(r"[，,]", admins_str) if a.strip()]
        rows.append((name, identifier, admin_names))
    wb.close()
    return rows


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, "导入表.xlsx")

    if CLIENT_ID == "your_client_id_here" or CLIENT_SECRET == "your_client_secret_here":
        print("错误: 请先修改脚本顶部的 CONFIG 区域，填写 CLIENT_ID 和 CLIENT_SECRET")
        sys.exit(1)

    print("=== PingCode 知识管理空间批量创建 ===\n")

    print(f"读取 Excel: {excel_path}")
    try:
        rows = read_excel(excel_path)
    except FileNotFoundError as e:
        print(f"错误: {e}")
        sys.exit(1)

    if not rows:
        print("Excel 中没有有效数据，退出。")
        sys.exit(0)

    print(f"共读取到 {len(rows)} 条有效记录\n")
    for i, (name, identifier, admins) in enumerate(rows):
        print(f"  [{i+1}] {name} ({identifier})  管理员: {', '.join(admins) if admins else '(无)'}")

    print("\n获取企业令牌...")
    try:
        token = get_enterprise_token()
        print("令牌获取成功\n")
    except Exception as e:
        print(f"获取令牌失败: {e}")
        sys.exit(1)

    success = 0
    fail = 0
    for i, (name, identifier, admins) in enumerate(rows):
        print(f"[{i+1}/{len(rows)}] 创建空间: {name} ({identifier})")

        try:
            space = create_wiki_space(token, name, identifier)
        except requests.HTTPError as e:
            print(f"  ✗ 创建失败: {e.response.status_code} {e.response.text[:200]}")
            fail += 1
            continue
        except Exception as e:
            print(f"  ✗ 创建失败: {e}")
            fail += 1
            continue

        space_id = space.get("id", "")
        if not space_id:
            print(f"  ✗ 响应中缺少空间 ID: {space}")
            fail += 1
            continue
        print(f"  ✓ 空间已创建, ID: {space_id}")

        if not admins:
            success += 1
            continue

        for admin_name in admins:
            try:
                user = find_user_by_name(token, admin_name)
            except Exception as e:
                print(f"  ✗ 查找用户 {admin_name} 失败: {e}")
                continue

            if not user:
                print(f"  ⚠ 未找到用户: {admin_name}")
                continue

            try:
                add_member(token, space_id, user["id"], ADMIN_ROLE_ID)
                print(f"  ✓ 已设置管理员: {user['display_name']} ({user['email']})")
            except Exception as e:
                print(f"  ✗ 添加管理员 {admin_name} 失败: {e}")

        success += 1

    print(f"\n=== 完成 ===\n成功: {success}  失败: {fail}")


if __name__ == "__main__":
    main()
