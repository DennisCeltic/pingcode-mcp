"""
PingCode 知识管理空间成员角色更新脚本

功能:
  1. 读取 pingcode/导入表.xlsx 获取每个空间的管理员名单
  2. 根据空间标识查询已创建的空间
  3. 删除现有成员，以管理员角色重新添加

使用方式:
  python3 pingcode/update_wiki_roles.py [--dry-run]

注意:
  PingCode wiki 成员 API 对 role_id 的支持可能有限制，
  如果更新后角色仍为"普通成员"，请手动在 PingCode UI 中设置管理员。
"""

import os
import re
import sys
import requests
from openpyxl import load_workbook

# ============================================================
# CONFIG
# ============================================================

CLIENT_ID = "ePrgDnMsGmZT"
CLIENT_SECRET = "SWrynTvzilneobmtruhuCHRI"
BASE_URL = "https://open.pingcode.com"
IS_PRIVATE = False
AUTH_TOKEN_PATH = "/v1/auth/token"
USERS_PATH = "/v1/directory/users"
WIKI_SPACES_PATH = "/v1/wiki/spaces"
WIKI_MEMBERS_PATH = "/v1/wiki/spaces/{space_id}/members"
WIKI_MEMBER_PATH = "/v1/wiki/spaces/{space_id}/members/{user_id}"
REQUEST_TIMEOUT = 30
ADMIN_ROLE_ID = "100000000000000000000001"

# ============================================================
# END CONFIG
# ============================================================


def api_path(path: str) -> str:
    prefix = "/open" if IS_PRIVATE else ""
    return f"{BASE_URL}{prefix}{path}"


def get_enterprise_token() -> str:
    url = api_path(AUTH_TOKEN_PATH)
    params = {
        "grant_type": "client_credentials",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
    }
    resp = requests.get(url, params=params, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    data = resp.json()
    access_token = data.get("access_token")
    if not access_token:
        raise RuntimeError(f"获取企业令牌失败，响应: {data}")
    return access_token


def find_user_by_name(token: str, name: str) -> dict | None:
    url = api_path(USERS_PATH)
    resp = requests.get(
        url,
        params={"keywords": name, "page_index": 0, "page_size": 10},
        headers={"Authorization": f"Bearer {token}"},
        timeout=REQUEST_TIMEOUT,
    )
    resp.raise_for_status()
    users = resp.json().get("values", [])
    if not users:
        return None
    for user in users:
        if user.get("display_name") == name.strip():
            return {"id": user["id"], "display_name": user["display_name"], "email": user.get("email", "")}
    user = users[0]
    return {"id": user["id"], "display_name": user["display_name"], "email": user.get("email", "")}


def find_space_by_identifier(token: str, identifier: str) -> dict | None:
    """根据空间标识查找已创建的空间"""
    url = api_path(WIKI_SPACES_PATH)
    page = 0
    while True:
        resp = requests.get(
            url,
            params={"page_index": page, "page_size": 50},
            headers={"Authorization": f"Bearer {token}"},
            timeout=REQUEST_TIMEOUT,
        )
        resp.raise_for_status()
        data = resp.json()
        for space in data.get("values", []):
            if space.get("identifier") == identifier:
                return space
        if len(data.get("values", [])) < 50:
            break
        page += 1
    return None


def get_space_members(token: str, space_id: str) -> list[dict]:
    """获取空间的所有成员"""
    url = api_path(WIKI_MEMBERS_PATH.format(space_id=space_id))
    all_members = []
    page = 0
    while True:
        resp = requests.get(
            url,
            params={"page_index": page, "page_size": 50},
            headers={"Authorization": f"Bearer {token}"},
            timeout=REQUEST_TIMEOUT,
        )
        resp.raise_for_status()
        members = resp.json().get("values", [])
        all_members.extend(members)
        if len(members) < 50:
            break
        page += 1
    return all_members


def delete_member(token: str, space_id: str, user_id: str) -> bool:
    url = api_path(WIKI_MEMBER_PATH.format(space_id=space_id, user_id=user_id))
    resp = requests.delete(url, headers={"Authorization": f"Bearer {token}"}, timeout=REQUEST_TIMEOUT)
    return resp.status_code == 200


def add_member(token: str, space_id: str, user_id: str, role_id: str = ADMIN_ROLE_ID) -> dict:
    url = api_path(WIKI_MEMBERS_PATH.format(space_id=space_id))
    body = {"member": {"id": user_id, "type": "user"}, "role_id": role_id}
    resp = requests.post(url, json=body, headers={"Authorization": f"Bearer {token}"}, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    return resp.json()


def read_excel(filepath: str) -> list[tuple[str, str, list[str]]]:
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"文件不存在: {filepath}")

    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        name = str(row[0]).strip() if row[0] is not None else ""
        identifier = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        admins_str = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""

        if not name or not identifier:
            continue
        if name in ("空间名称", "Name") and identifier in ("空间标识", "Identifier"):
            continue

        admin_names = [a.strip() for a in re.split(r"[，,]", admins_str) if a.strip()]
        rows.append((name, identifier, admin_names))
    wb.close()
    return rows


def main():
    dry_run = "--dry-run" in sys.argv

    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, "导入表.xlsx")

    if dry_run:
        print("*** DRY RUN 模式 - 不会实际修改任何数据 ***\n")

    print("=== PingCode 知识空间成员角色更新 ===\n")

    print(f"读取 Excel: {excel_path}")
    try:
        rows = read_excel(excel_path)
    except FileNotFoundError as e:
        print(f"错误: {e}")
        sys.exit(1)

    if not rows:
        print("Excel 中没有有效数据，退出。")
        sys.exit(0)

    print(f"共 {len(rows)} 条记录\n")

    print("获取企业令牌...")
    try:
        token = get_enterprise_token()
        print("令牌获取成功\n")
    except Exception as e:
        print(f"获取令牌失败: {e}")
        sys.exit(1)

    # 预加载所有用户缓存
    print("预加载用户信息...")
    user_cache = {}
    all_admin_names = set()
    for _, _, admins in rows:
        all_admin_names.update(admins)
    for name in all_admin_names:
        user = find_user_by_name(token, name)
        if user:
            user_cache[name] = user
            print(f"  ✓ {name} → {user['id']}")
        else:
            print(f"  ⚠ 未找到用户: {name}")
    print()

    for i, (name, identifier, admin_names) in enumerate(rows):
        print(f"[{i+1}/{len(rows)}] {name} ({identifier})")

        space = find_space_by_identifier(token, identifier)
        if not space:
            print(f"  ⚠ 未找到空间 (identifier={identifier})，跳过")
            continue

        space_id = space["id"]
        print(f"  空间 ID: {space_id}")

        current_members = get_space_members(token, space_id)
        print(f"  当前成员: {len(current_members)} 人")
        for m in current_members:
            role_name = m.get("role", {}).get("name", "?")
            user_name = m.get("user", {}).get("display_name", "?")
            print(f"    - {user_name} ({role_name})")

        expected_user_ids = []
        for admin_name in admin_names:
            user = user_cache.get(admin_name)
            if user:
                expected_user_ids.append(user["id"])
            else:
                print(f"  ⚠ 跳过 {admin_name}（用户未找到）")

        if not expected_user_ids:
            print(f"  ⚠ 没有有效的管理员用户，跳过")
            continue

        # Step 1: 删除现有成员
        print(f"\n  更新成员...")
        for m in current_members:
            m_user_id = m.get("user", {}).get("id", "")
            m_user_name = m.get("user", {}).get("display_name", "?")
            if m_user_id:
                if dry_run:
                    print(f"    [DRY-RUN] 将删除: {m_user_name}")
                else:
                    ok = delete_member(token, space_id, m_user_id)
                    status = "✓" if ok else "✗"
                    print(f"    {status} 已删除: {m_user_name}")

        # Step 2: 重新添加为管理员
        for uid in expected_user_ids:
            user_info = next((u for u in user_cache.values() if u["id"] == uid), None)
            display = user_info["display_name"] if user_info else uid
            if dry_run:
                print(f"    [DRY-RUN] 将添加: {display} (role_id={ADMIN_ROLE_ID})")
            else:
                try:
                    result = add_member(token, space_id, uid, ADMIN_ROLE_ID)
                    actual_role = result.get("role", {}).get("name", "?")
                    print(f"    ✓ 已添加: {display} → {actual_role}")
                    if actual_role != "管理员":
                        print(f"      ⚠ 角色仍为「{actual_role}」，请在 PingCode UI 中手动设置为管理员")
                except Exception as e:
                    print(f"    ✗ 添加 {display} 失败: {e}")

        print()

    print("=== 完成 ===")
    if dry_run:
        print("使用 'python3 pingcode/update_wiki_roles.py' 执行实际更新")


if __name__ == "__main__":
    main()
